from pathlib import Path
from typing import Tuple

import openpyxl
import pandas as pd

from .models import (
    TgaMetadata,
)

__all__ = ("parse_tga_xlsx",)


def parse_tga_metadata(path: Path) -> Tuple[int, TgaMetadata]:
    if not path.exists():
        raise RuntimeError(f"Provided path does not exist: {path!r}")

    wb = openpyxl.load_workbook(path)

    # the workbooks should only have a single sheet
    ws = wb[wb.sheetnames[0]]

    data_start, temperature_program_start, temperature_program_end = 0, 0, 0
    metadata = {}
    for i, row in enumerate(ws.iter_rows()):
        i = i + 1  # excel and openpyxl uses 1-indexing
        row = ws[i]
        if row[0].value == "Time":
            data_start = i
            break
        elif row[0].value == "Temperature Program":
            temperature_program_start = i
            continue
        elif row[1].value == "Temperature Program Mode":
            temperature_program_end = i
            metadata["Temperature Program Mode"] = row[2].value
            continue
        elif row[0].value == "Comment":
            comment_start = i
        elif not row[0].value:
            continue

        else:
            metadata[row[0].value] = row[1].value
    if data_start == 0:
        raise RuntimeError("Unexpected format for TGA-DTA xlsx file: no data header found")
    if temperature_program_start == 0:
        raise RuntimeError("Unexpected format for TGA-DTA xlsx file: no temperature program found")
    if temperature_program_end == 0:
        raise RuntimeError(
            "Unexpected format for TGA-DTA xlsx file: no end of temperature program found"
        )

    temperature_program_steps = []

    row = ws[temperature_program_start]
    temperature_program_units = {
        "initial_temp": row[2].value,
        "final_temp": row[3].value,
        "ramp_rate": row[4].value,
        "hold_time": row[5].value,
        "sampling_time": row[6].value,
    }

    for i in range(temperature_program_start + 1, temperature_program_end):
        row = ws[i]
        step = {
            "initial_temp": row[2].value,
            "final_temp": row[3].value,
            "ramp_rate": row[4].value,
            "hold_time": row[5].value,
            "sampling_time": row[6].value,
            "gas1_status": row[7].value,
            "gas2_status": row[8].value,
            "store": row[9].value,
        }
        temperature_program_steps.append(step)

    metadata["temperature_program"] = {
        "steps": temperature_program_steps,
        "units": temperature_program_units,
    }

    for row in ws[comment_start:data_start]:
        if not row[1].value or ":" not in row[1].value:
            continue
        split = row[1].value.split(":")
        key = split[0]
        val = ":".join(split[1:])
        metadata[key] = val

    return data_start, TgaMetadata(**metadata)


def parse_tga_xlsx(path: Path) -> Tuple[pd.DataFrame, TgaMetadata]:
    """Parses a excel worksheet containing tga/dta data from a XX thermal anaylsis instrument
    returns a tuple of (data, metadata).
    """

    if str == str:
        path = Path(path)

    data_start, metadata = parse_tga_metadata(path)

    df = pd.read_excel(path, header=data_start - 1, skiprows=lambda x: x == data_start)
    # breakpoint()
    df.rename(
        columns={
            "Time": "elapsed time (min)",
            "Temp.": "temperature (°C)",
            "DTA": "DTA voltage (μV)",
            "TG": "mass (μg)",
            "DTG": "differential mass (μg/min)",
        },
        inplace=True,
    )

    df.dropna(how="all", axis=1, inplace=True)  # drop empty columns

    # data = TgaData(tabbular_data=df)

    return df, metadata
